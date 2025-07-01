import pandas as pd
import openpyxl
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import io
from datetime import datetime
from typing import Dict, List, Any, Optional
import logging
from components.database_manager import DatabaseManager
from components.query_processor import QueryProcessor

class ReportGenerator:
    """Excel 報表生成器"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.db_manager = DatabaseManager()
        self.query_processor = QueryProcessor()
        
        # 定義樣式
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_status_report(self) -> Optional[bytes]:
        """生成配件狀態報表"""
        try:
            workbook = openpyxl.Workbook()
            
            # 移除預設工作表
            workbook.remove(workbook.active)
            
            # 建立各個工作表
            self._create_overview_sheet(workbook)
            self._create_pat_status_sheet(workbook)
            self._create_kyec_status_sheet(workbook)
            self._create_customer_analysis_sheet(workbook)
            self._create_charts_sheet(workbook)
            
            return self._save_workbook_to_bytes(workbook)
            
        except Exception as e:
            self.logger.error(f"狀態報表生成失敗: {str(e)}")
            return None
    
    def _create_overview_sheet(self, workbook):
        """建立總覽工作表"""
        ws = workbook.create_sheet("總覽統計")
        
        # 標題
        ws['A1'] = "配件狀態總覽報表"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # 獲取統計資料
        overview_stats = self.db_manager.get_overview_statistics()
        
        # 總覽統計
        ws['A4'] = "總覽統計"
        ws['A4'].font = self.header_font
        ws['A4'].fill = self.header_fill
        
        stats_data = [
            ["項目", "數量"],
            ["配件總數", overview_stats.get('total_parts', 0)],
            ["維修中配件", overview_stats.get('repair_parts', 0)],
            ["正常生產配件", overview_stats.get('normal_parts', 0)],
            ["客戶借出配件", overview_stats.get('borrowed_parts', 0)]
        ]
        
        for row_idx, row_data in enumerate(stats_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 5:  # 標題行
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
        
        # PAT 狀態分佈
        pat_status = self.db_manager.get_pat_status_distribution()
        if not pat_status.empty:
            ws['A11'] = "PAT 配件狀態分佈"
            ws['A11'].font = self.header_font
            ws['A11'].fill = self.header_fill
            
            self._write_dataframe_to_sheet(ws, pat_status, start_row=12)
        
        # KYEC 狀態分佈
        kyec_status = self.db_manager.get_kyec_status_distribution()
        if not kyec_status.empty:
            ws['D11'] = "KYEC 配件狀態分佈"
            ws['D11'].font = self.header_font
            ws['D11'].fill = self.header_fill
            
            self._write_dataframe_to_sheet(ws, kyec_status, start_row=12, start_col=4)
        
        # 調整欄寬
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_pat_status_sheet(self, workbook):
        """建立 PAT 狀態詳細工作表"""
        ws = workbook.create_sheet("PAT 配件詳細")
        
        # 標題
        ws['A1'] = "PAT 配件狀態詳細分析"
        ws['A1'].font = Font(size=14, bold=True)
        
        # 獲取詳細統計
        pat_stats = self.db_manager.get_detailed_pat_statistics()
        
        if not pat_stats.empty:
            self._write_dataframe_to_sheet(ws, pat_stats, start_row=3, with_table=True)
            
            # 添加條件格式化
            self._apply_conditional_formatting(ws, pat_stats, start_row=4)
    
    def _create_kyec_status_sheet(self, workbook):
        """建立 KYEC 狀態詳細工作表"""
        ws = workbook.create_sheet("KYEC 配件詳細")
        
        # 標題
        ws['A1'] = "KYEC 配件狀態詳細分析"
        ws['A1'].font = Font(size=14, bold=True)
        
        # 獲取詳細統計
        kyec_stats = self.db_manager.get_detailed_kyec_statistics()
        
        if not kyec_stats.empty:
            self._write_dataframe_to_sheet(ws, kyec_stats, start_row=3, with_table=True)
            
            # 添加條件格式化
            self._apply_conditional_formatting(ws, kyec_stats, start_row=4)
    
    def _create_customer_analysis_sheet(self, workbook):
        """建立客戶分析工作表"""
        ws = workbook.create_sheet("客戶分析")
        
        # 標題
        ws['A1'] = "客戶配件分析報表"
        ws['A1'].font = Font(size=14, bold=True)
        
        # 獲取客戶統計
        customer_stats = self.db_manager.get_customer_statistics()
        
        if not customer_stats.empty:
            self._write_dataframe_to_sheet(ws, customer_stats, start_row=3, with_table=True)
            
            # 添加客戶排行榜
            top_customers = customer_stats.head(10)
            if not top_customers.empty:
                ws['A' + str(len(customer_stats) + 8)] = "客戶配件數量排行榜 (前10名)"
                ws['A' + str(len(customer_stats) + 8)].font = Font(size=12, bold=True)
                
                self._write_dataframe_to_sheet(
                    ws, top_customers, 
                    start_row=len(customer_stats) + 10, 
                    with_table=True
                )
    
    def _create_charts_sheet(self, workbook):
        """建立圖表工作表"""
        ws = workbook.create_sheet("圖表分析")
        
        # 標題
        ws['A1'] = "配件狀態圖表分析"
        ws['A1'].font = Font(size=14, bold=True)
        
        # 獲取資料用於圖表
        pat_status = self.db_manager.get_pat_status_distribution()
        kyec_status = self.db_manager.get_kyec_status_distribution()
        
        # PAT 狀態圓餅圖
        if not pat_status.empty:
            self._add_pie_chart(ws, pat_status, "PAT 配件狀態分佈", "A3", "E15")
        
        # KYEC 狀態圓餅圖
        if not kyec_status.empty:
            self._add_pie_chart(ws, kyec_status, "KYEC 配件狀態分佈", "G3", "K15")
    
    def _write_dataframe_to_sheet(self, worksheet, df, start_row=1, start_col=1, with_table=False):
        """將 DataFrame 寫入工作表"""
        # 寫入標題
        for col_idx, column_name in enumerate(df.columns, start=start_col):
            cell = worksheet.cell(row=start_row, column=col_idx, value=column_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # 寫入資料
        for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row, start=start_col):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
        
        # 建立表格
        if with_table:
            end_row = start_row + len(df)
            end_col = start_col + len(df.columns) - 1
            
            table_range = f"{worksheet.cell(start_row, start_col).coordinate}:{worksheet.cell(end_row, end_col).coordinate}"
            table = Table(displayName=f"Table{start_row}", ref=table_range)
            
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True
            )
            table.tableStyleInfo = style
            worksheet.add_table(table)
    
    def _add_pie_chart(self, worksheet, data, title, data_start, chart_position):
        """添加圓餅圖"""
        # 先將資料寫入工作表
        start_row = int(data_start[1:])
        start_col = ord(data_start[0]) - ord('A') + 1
        
        self._write_dataframe_to_sheet(worksheet, data, start_row, start_col)
        
        # 建立圓餅圖
        chart = PieChart()
        chart.title = title
        
        # 設定資料範圍
        data_range = Reference(
            worksheet,
            min_col=start_col + 1,
            min_row=start_row + 1,
            max_row=start_row + len(data)
        )
        
        categories = Reference(
            worksheet,
            min_col=start_col,
            min_row=start_row + 1,
            max_row=start_row + len(data)
        )
        
        chart.add_data(data_range, titles_from_data=False)
        chart.set_categories(categories)
        
        # 添加圖表到工作表
        worksheet.add_chart(chart, chart_position)
    
    def _apply_conditional_formatting(self, worksheet, df, start_row):
        """套用條件格式化"""
        # 為維修相關的欄位添加紅色背景
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        
        for row_idx in range(start_row, start_row + len(df)):
            for col_idx, column_name in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                if "維修" in column_name and isinstance(cell.value, (int, float)) and cell.value > 0:
                    cell.fill = red_fill
                elif "正常" in column_name and isinstance(cell.value, (int, float)) and cell.value > 0:
                    cell.fill = green_fill
    
    def generate_trend_report(self, date_range=None) -> Optional[bytes]:
        """生成趨勢分析報表"""
        try:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
            
            # 建立趨勢分析工作表
            ws = workbook.create_sheet("趨勢分析")
            
            # 標題
            ws['A1'] = "配件狀態趨勢分析報表"
            ws['A1'].font = Font(size=16, bold=True)
            
            if date_range:
                ws['A2'] = f"分析期間: {date_range[0]} 至 {date_range[1]}"
            
            # 獲取趨勢資料
            trend_data = self.db_manager.get_status_trend_data()
            
            if not trend_data.empty:
                self._write_dataframe_to_sheet(ws, trend_data, start_row=4, with_table=True)
                
                # 添加趨勢圖表
                self._add_trend_chart(ws, trend_data)
            
            # 維修週期分析
            maintenance_data = self.db_manager.get_maintenance_cycle_data()
            if not maintenance_data.empty:
                ws2 = workbook.create_sheet("維修週期分析")
                ws2['A1'] = "維修週期分析"
                ws2['A1'].font = Font(size=14, bold=True)
                
                self._write_dataframe_to_sheet(ws2, maintenance_data, start_row=3, with_table=True)
            
            return self._save_workbook_to_bytes(workbook)
            
        except Exception as e:
            self.logger.error(f"趨勢報表生成失敗: {str(e)}")
            return None
    
    def _add_trend_chart(self, worksheet, data):
        """添加趨勢圖表"""
        try:
            chart = LineChart()
            chart.title = "配件狀態變化趨勢"
            chart.style = 13
            chart.x_axis.title = "時間"
            chart.y_axis.title = "數量"
            
            # 設定資料範圍（這裡需要根據實際資料結構調整）
            data_range = Reference(
                worksheet,
                min_col=2,
                min_row=4,
                max_col=len(data.columns),
                max_row=4 + len(data)
            )
            
            chart.add_data(data_range, titles_from_data=True)
            
            # 設定類別軸
            categories = Reference(
                worksheet,
                min_col=1,
                min_row=5,
                max_row=4 + len(data)
            )
            chart.set_categories(categories)
            
            worksheet.add_chart(chart, "A20")
            
        except Exception as e:
            self.logger.warning(f"趨勢圖表添加失敗: {str(e)}")
    
    def generate_change_log_report(self, days=30) -> Optional[bytes]:
        """生成變更記錄報表"""
        try:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
            
            # 建立變更記錄工作表
            ws = workbook.create_sheet("變更記錄")
            
            # 標題
            ws['A1'] = f"配件變更記錄報表 (最近 {days} 天)"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f"生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # 獲取變更記錄
            change_logs = self.db_manager.get_change_logs(days_filter=f"最近{days}天")
            
            if not change_logs.empty:
                # 選擇重要欄位
                important_columns = ['timestamp', 'table_name', 'operation', 'row_key', 
                                   'column_name', 'old_value', 'new_value', 'user']
                
                display_data = change_logs[important_columns].copy()
                display_data.columns = ['變更時間', '資料表', '操作類型', '配件識別', 
                                      '變更欄位', '原值', '新值', '操作人員']
                
                self._write_dataframe_to_sheet(ws, display_data, start_row=4, with_table=True)
            
            # 變更統計分析
            ws2 = workbook.create_sheet("變更統計")
            ws2['A1'] = "變更統計分析"
            ws2['A1'].font = Font(size=14, bold=True)
            
            # 按日期統計
            if not change_logs.empty:
                daily_stats = change_logs.groupby(change_logs['timestamp'].dt.date).size().reset_index()
                daily_stats.columns = ['日期', '變更次數']
                
                self._write_dataframe_to_sheet(ws2, daily_stats, start_row=3, with_table=True)
                
                # 按操作類型統計
                operation_stats = change_logs['operation'].value_counts().reset_index()
                operation_stats.columns = ['操作類型', '次數']
                
                ws2['D1'] = "操作類型統計"
                ws2['D1'].font = Font(size=12, bold=True)
                self._write_dataframe_to_sheet(ws2, operation_stats, start_row=3, start_col=4, with_table=True)
            
            return self._save_workbook_to_bytes(workbook)
            
        except Exception as e:
            self.logger.error(f"變更記錄報表生成失敗: {str(e)}")
            return None
    
    def generate_custom_report(self, report_name: str, tables: List[str], 
                             columns: List[str], filter_conditions: str = "") -> Optional[bytes]:
        """生成自訂報表"""
        try:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
            
            ws = workbook.create_sheet(report_name)
            
            # 標題
            ws['A1'] = report_name
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f"生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # 建立查詢
            for i, table in enumerate(tables):
                # 建立 SQL 查詢
                selected_columns = [col for col in columns if col in self.db_manager.get_available_columns([table])]
                
                if selected_columns:
                    query = f"SELECT {', '.join(selected_columns)} FROM {table}"
                    
                    if filter_conditions:
                        query += f" WHERE {filter_conditions}"
                    
                    query += " LIMIT 1000"  # 限制結果數量
                    
                    # 執行查詢
                    if self.db_manager.validate_sql_query(query):
                        result = self.db_manager.execute_query(query)
                        
                        if not result.empty:
                            # 計算起始行
                            start_row = 4 + i * (len(result) + 5)
                            
                            # 添加表格標題
                            ws[f'A{start_row - 1}'] = f"{table} 資料"
                            ws[f'A{start_row - 1}'].font = Font(size=12, bold=True)
                            
                            # 寫入資料
                            self._write_dataframe_to_sheet(ws, result, start_row=start_row, with_table=True)
            
            return self._save_workbook_to_bytes(workbook)
            
        except Exception as e:
            self.logger.error(f"自訂報表生成失敗: {str(e)}")
            return None
    
    def _save_workbook_to_bytes(self, workbook) -> bytes:
        """將工作簿儲存為位元組"""
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_comprehensive_report(self) -> Optional[bytes]:
        """生成綜合報表"""
        try:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
            
            # 執行摘要
            self._create_executive_summary(workbook)
            
            # 詳細分析
            self._create_overview_sheet(workbook)
            self._create_pat_status_sheet(workbook)
            self._create_kyec_status_sheet(workbook)
            self._create_customer_analysis_sheet(workbook)
            
            # 趨勢和變更分析
            self._create_trend_analysis_sheet(workbook)
            self._create_change_analysis_sheet(workbook)
            
            # 圖表總覽
            self._create_charts_sheet(workbook)
            
            return self._save_workbook_to_bytes(workbook)
            
        except Exception as e:
            self.logger.error(f"綜合報表生成失敗: {str(e)}")
            return None
    
    def _create_executive_summary(self, workbook):
        """建立執行摘要"""
        ws = workbook.create_sheet("執行摘要")
        
        # 標題
        ws['A1'] = "配件管理執行摘要"
        ws['A1'].font = Font(size=18, bold=True)
        ws['A2'] = f"報表日期: {datetime.now().strftime('%Y年%m月%d日')}"
        
        # 關鍵指標
        overview_stats = self.db_manager.get_overview_statistics()
        
        ws['A4'] = "關鍵績效指標"
        ws['A4'].font = Font(size=14, bold=True)
        
        kpi_data = [
            ["指標", "數值", "說明"],
            ["配件總數", overview_stats.get('total_parts', 0), "系統中所有配件的總數量"],
            ["可用率", f"{(overview_stats.get('normal_parts', 0) / max(overview_stats.get('total_parts', 1), 1) * 100):.1f}%", "正常生產配件佔總數的百分比"],
            ["維修率", f"{(overview_stats.get('repair_parts', 0) / max(overview_stats.get('total_parts', 1), 1) * 100):.1f}%", "維修中配件佔總數的百分比"],
            ["借出率", f"{(overview_stats.get('borrowed_parts', 0) / max(overview_stats.get('total_parts', 1), 1) * 100):.1f}%", "客戶借出配件佔總數的百分比"]
        ]
        
        for row_idx, row_data in enumerate(kpi_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 5:  # 標題行
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
        
        # 調整欄寬
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
    
    def _create_trend_analysis_sheet(self, workbook):
        """建立趨勢分析工作表"""
        ws = workbook.create_sheet("趨勢分析")
        
        ws['A1'] = "配件狀態趨勢分析"
        ws['A1'].font = Font(size=14, bold=True)
        
        trend_data = self.db_manager.get_status_trend_data()
        if not trend_data.empty:
            self._write_dataframe_to_sheet(ws, trend_data, start_row=3, with_table=True)
    
    def _create_change_analysis_sheet(self, workbook):
        """建立變更分析工作表"""
        ws = workbook.create_sheet("變更分析")
        
        ws['A1'] = "配件變更分析"
        ws['A1'].font = Font(size=14, bold=True)
        
        # 最近30天的變更統計
        change_logs = self.db_manager.get_change_logs(days_filter="最近30天")
        
        if not change_logs.empty:
            # 按日期統計變更
            daily_changes = change_logs.groupby(change_logs['timestamp'].dt.date).size().reset_index()
            daily_changes.columns = ['日期', '變更次數']
            
            self._write_dataframe_to_sheet(ws, daily_changes, start_row=3, with_table=True)